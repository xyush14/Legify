"""Word / Excel attachment support.

Lawyers routinely keep drafts, notices, and cause-lists as Word (.docx) and
Excel (.xlsx) files. Everywhere we accept a PDF or a photo we now also accept
these — but they are handled on a DIFFERENT path from images: an office file is
already-digital text, so we extract its text directly (python-docx / openpyxl)
instead of sending it through the Groq vision OCR pipeline (which only reads
images). The extracted text then feeds the same downstream flow as an OCR
transcription — draft prefill, field extraction, or the searchable vault.

Scope is the modern XML formats (.docx / .xlsx, Office 2007+). The old binary
.doc / .xls are recognised only so we can return a clear "please save as
.docx/.xlsx" message rather than a cryptic parse error — reading them needs
LibreOffice/antiword on the server, which we deliberately don't ship.
"""

from __future__ import annotations

import io
import logging
import os
from typing import Optional

log = logging.getLogger(__name__)

# Modern (supported) Office Open XML formats → canonical kind.
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Legacy binary formats — recognised only to give a friendly error.
_DOC_MIME = "application/msword"
_XLS_MIME = "application/vnd.ms-excel"

# The MIME allow-list callers add to their existing image/PDF set.
OFFICE_ALLOWED_MIME = {_DOCX_MIME, _XLSX_MIME, _DOC_MIME, _XLS_MIME}

# The image/PDF formats that take the vision-OCR path (everything that isn't an
# office file). Kept here so upload validation lives in one place.
MEDIA_MIME = {"image/jpeg", "image/png", "image/webp", "image/gif", "application/pdf"}

# Browsers (and WhatsApp) sometimes send a generic or empty content-type for
# attachments, so we also key off the filename extension.
_EXT_KIND = {
    ".docx": "docx", ".xlsx": "xlsx",
    ".doc": "legacy-doc", ".xls": "legacy-xls",
}
_MIME_KIND = {
    _DOCX_MIME: "docx", _XLSX_MIME: "xlsx",
    _DOC_MIME: "legacy-doc", _XLS_MIME: "legacy-xls",
}

# The <input accept="..."> fragment to append to the existing image/PDF accept.
# Includes both the extensions (robust across OSes) and the canonical MIME types.
ACCEPT_OFFICE = (
    ".doc,.docx,.xls,.xlsx"
    f",{_DOCX_MIME},{_XLSX_MIME},{_DOC_MIME},{_XLS_MIME}"
)


def office_kind(mime: Optional[str], filename: Optional[str]) -> Optional[str]:
    """Classify an upload as a Word/Excel file.

    Returns "docx" | "xlsx" | "legacy-doc" | "legacy-xls", or None if this is
    not an office file (i.e. it's an image/PDF and should take the OCR path).
    Prefers the MIME type but falls back to the filename extension, since
    browsers are inconsistent about the content-type they attach.
    """
    mt = (mime or "").split(";")[0].strip().lower()
    if mt in _MIME_KIND:
        return _MIME_KIND[mt]
    name = (filename or "").lower()
    _, ext = os.path.splitext(name)
    return _EXT_KIND.get(ext)


def is_office_upload(mime: Optional[str], filename: Optional[str]) -> bool:
    return office_kind(mime, filename) is not None


class UnsupportedOfficeFormat(ValueError):
    """A recognised-but-unreadable office file (legacy .doc / .xls)."""


def extract_office_text(data: bytes, mime: Optional[str], filename: Optional[str]) -> str:
    """Extract readable text from a Word/Excel upload.

    .docx → paragraphs + tables (tables rendered as Markdown) in document order.
    .xlsx → each sheet rendered as a Markdown table.
    Legacy .doc/.xls raise UnsupportedOfficeFormat with a user-facing message.
    Raises ValueError on a corrupt/unreadable file.
    """
    kind = office_kind(mime, filename)
    if kind == "docx":
        return _extract_docx(data)
    if kind == "xlsx":
        return _extract_xlsx(data)
    if kind == "legacy-doc":
        raise UnsupportedOfficeFormat(
            "old .doc files aren't supported — open it in Word and 'Save As' .docx, then upload again"
        )
    if kind == "legacy-xls":
        raise UnsupportedOfficeFormat(
            "old .xls files aren't supported — open it in Excel and 'Save As' .xlsx, then upload again"
        )
    raise ValueError("not a Word or Excel file")


def collect_uploads(entries, *, max_bytes: int):
    """Validate & split a batch of uploads into OCR pages and office text.

    `entries` is a list of (data: bytes, content_type: str, filename: str) that
    the caller has already read off the wire. Each item is either an image/PDF
    (kept for the vision-OCR path) or a Word/Excel file (text extracted here and
    concatenated). Returns (media_pages, office_text).

    Raises ValueError with a page-scoped, user-facing message on an empty/oversize
    file, an unsupported type, or an unreadable office file — the caller maps that
    to its own 400 response style.
    """
    media: list[tuple[bytes, str]] = []
    texts: list[str] = []
    mb = max(1, max_bytes // (1024 * 1024))
    for idx, (data, mime, filename) in enumerate(entries, start=1):
        if not data:
            raise ValueError(f"page {idx}: empty file")
        if len(data) > max_bytes:
            raise ValueError(f"page {idx}: too large; max {mb} MB")
        if office_kind(mime, filename) is not None:
            try:
                texts.append(extract_office_text(data, mime, filename))
            except ValueError as e:
                raise ValueError(f"page {idx}: {e}")
            continue
        if (mime or "").split(";")[0].strip().lower() not in MEDIA_MIME:
            raise ValueError(
                f"page {idx}: unsupported file type {mime!r}; "
                "use an image, PDF, Word (.docx) or Excel (.xlsx) file"
            )
        media.append((data, mime))
    return media, "\n\n".join(t for t in texts if t.strip())


def _extract_docx(data: bytes) -> str:
    """Read a .docx into text, keeping paragraphs and tables in reading order."""
    try:
        from docx import Document
        from docx.document import Document as _Doc
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as e:  # pragma: no cover - dependency is pinned
        raise ValueError(f"cannot read .docx (python-docx missing): {e}")

    try:
        doc = Document(io.BytesIO(data))
    except Exception as e:  # noqa: BLE001 - corrupt/not-a-docx
        raise ValueError(f"couldn't open the Word file — is it a valid .docx? ({e})")

    def _iter_block_items(parent):
        # Walk the body in document order so paragraphs and tables interleave
        # the way a reader sees them (python-docx exposes them as separate lists).
        body = parent.element.body
        for child in body.iterchildren():
            tag = child.tag.split("}")[-1]
            if tag == "p":
                yield Paragraph(child, parent)
            elif tag == "tbl":
                yield Table(child, parent)

    out: list[str] = []
    for block in _iter_block_items(doc):
        if isinstance(block, Paragraph):
            txt = block.text.strip()
            if txt:
                out.append(txt)
        elif isinstance(block, Table):
            md = _table_to_markdown([[c.text.strip() for c in row.cells] for row in block.rows])
            if md:
                out.append(md)
    return "\n\n".join(out).strip()


def _extract_xlsx(data: bytes) -> str:
    """Read a .xlsx into text — one Markdown table per non-empty sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - dependency is pinned
        raise ValueError(f"cannot read .xlsx (openpyxl missing): {e}")

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - corrupt/not-a-xlsx
        raise ValueError(f"couldn't open the Excel file — is it a valid .xlsx? ({e})")

    sections: list[str] = []
    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if any(cells):
                    rows.append(cells)
            if not rows:
                continue
            md = _table_to_markdown(rows)
            title = (ws.title or "Sheet").strip()
            sections.append(f"## {title}\n\n{md}" if len(wb.worksheets) > 1 else md)
    finally:
        wb.close()
    return "\n\n".join(sections).strip()


def _table_to_markdown(rows: list[list[str]]) -> str:
    """Render a rectangular list of rows as a GitHub-flavoured Markdown table.

    The first row is treated as the header. Cells are padded to a uniform
    column count so ragged rows still line up.
    """
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return ""
    ncols = max(len(r) for r in rows)
    norm = [[_md_cell(r[i]) if i < len(r) else "" for i in range(ncols)] for r in rows]
    header, body = norm[0], norm[1:]
    lines = ["| " + " | ".join(header) + " |",
             "| " + " | ".join(["---"] * ncols) + " |"]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _md_cell(v: str) -> str:
    """Escape pipes/newlines so a cell can't break the Markdown table grid."""
    return (v or "").replace("\n", " ").replace("|", "\\|").strip()
